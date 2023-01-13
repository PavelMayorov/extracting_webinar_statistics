import os
import logging

from dotenv import load_dotenv

from openpyxl.styles import PatternFill, Alignment


# Создание и настройка логгера
def init_logger(name: str) -> logging.Logger:
    logger = logging.getLogger(name)
    logger.setLevel(logging.DEBUG)
    sh = logging.StreamHandler()
    sh.setLevel(logging.DEBUG)
    FORMAT = '%(asctime)s :: %(name)s :: %(lineno)s - %(levelname)s - %(message)s'
    sh.setFormatter(logging.Formatter(FORMAT))
    logger.addHandler(sh)
    logging.getLogger('urllib3').setLevel('ERROR')
    return logger


logger = init_logger('EWS')

# Настройки ячеек excel
RED_FILL = PatternFill(start_color="FF1E3A", end_color="FF1E3A", fill_type="solid")
GREEN_FILL = PatternFill(start_color="7CFC00", end_color="7CFC00", fill_type="solid")
CENTERED_TEXT = Alignment(horizontal='center')

# Заголовки для обращения к API Webinar
load_dotenv()
HEADERS = {
    'x-auth-token': os.getenv('WEBINAR_TOKEN'),
    'Content-Type': 'application/x-www-form-urlencoded',
}
