import typing as tp

import openpyxl

from settings import RED_FILL, GREEN_FILL, CENTERED_TEXT, logger


def cell_entry(sheet, row, column, value, fill=None, alignment=CENTERED_TEXT):
    cell = sheet.cell(row=row, column=column)
    cell.value = value
    if fill:
        cell.fill = fill
    cell.alignment = alignment


def write_users_info(data: tp.List[dict], file_name: str) -> None:
    data = {d['email']: d['name'] for d in data}
    students_list = openpyxl.Workbook()
    sheet = students_list.create_sheet('Лист1', 0)
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 40
    for i, (email, name) in enumerate(data.items()):
        cell_entry(sheet, i + 3, 1, i + 1)
        cell_entry(sheet, i + 3, 3, name)
        cell_entry(sheet, i + 3, 4, email)
    students_list.save(file_name)


def write_statistics(events_data: tp.List[dict], file_name: str) -> None:
    try:
        students_list = openpyxl.load_workbook(filename=file_name)
    except Exception as ex:
        logger.warning(f'Не удалось открыть файл! Проверьте имя файла и его расположение! \n{ex}')
    else:
        sheet = students_list['Лист1']
        row = 1
        column = 4
        for event_data in events_data:
            date = event_data['date']
            if sheet.cell(row=row, column=column).value != date:
                column += 1
                sheet.cell(row=row, column=column).value = date
                sheet.column_dimensions[chr(64+column)].width = 10
            i = 2
            while (sheet.cell(row=row+i, column=4).value or sheet.cell(row=row+i+1, column=4).value or
                   sheet.cell(row=row+i+2, column=4).value):
                if not sheet.cell(row=row+i, column=column).value and sheet.cell(row=row+i, column=4).value:
                    cell_entry(sheet, row+i, column, 'н', RED_FILL)
                if event_data['email'].lower() == str(sheet.cell(row=row+i, column=4).value).lower().strip():
                    if event_data['actual_presence'] == '100.00' or event_data['actual_presence'] >= '60.00':
                        cell_entry(sheet, row+i, column, event_data['actual_presence'], GREEN_FILL)
                    else:
                        cell_entry(sheet, row+i, column, event_data['actual_presence'], RED_FILL)
                i += 1
        students_list.save(file_name)
        logger.info(f'Данные о мероприятиях успешно сохранены в файл {file_name}.')
